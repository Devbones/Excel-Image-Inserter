import tkinter as tk
from tkinter import filedialog, ttk
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage
import os
import threading
from queue import Queue
import shutil

# Constants for image compression
QUALITY = 25  # Compression quality (lower means more compression)
TOTAL_ORIGINAL = 0
TOTAL_COMPRESSED = 0
TOTAL_GAIN = 0
TOTAL_FILES = 0
CONVERT_PNG_TO_JPG = False  # Modify this if PNG to JPG conversion is required


# Function to resize the image maintaining aspect ratio to fit cell size
def resize_image_to_fit_cell(image_path, cell_width, cell_height):
    with PILImage.open(image_path) as img:
        img_width, img_height = img.size
        aspect_ratio = img_width / img_height

        # Calculate new dimensions to fit the cell while maintaining aspect ratio
        if cell_width / cell_height > aspect_ratio:
            new_height = int(cell_height)
            new_width = int(new_height * aspect_ratio)
        else:
            new_width = int(cell_width)
            new_height = int(new_width / aspect_ratio)

        resized_img = img.resize((new_width, new_height), PILImage.Resampling.LANCZOS)
        return resized_img  # Return resized image object


# Function to compress images in a directory
def compress(input_dir, output_dir, progress_label, progress_bar):
    global TOTAL_ORIGINAL, TOTAL_COMPRESSED, TOTAL_GAIN, TOTAL_FILES

    # Get total image files for compression progress calculation
    total_files = sum([1 for _, _, files in os.walk(input_dir) for file in files if file.lower().endswith(('.png', '.jpg', '.jpeg', '.tiff', '.bmp', '.gif', '.webp'))])

    compressed_files = 0  # Track how many files have been compressed

    for root, dirs, files in os.walk(input_dir):
        for file in files:
            input_path = os.path.join(root, file)
            rel_path = os.path.relpath(input_path, input_dir)
            output_path = os.path.join(output_dir, rel_path)

            if file.lower().endswith(('.png', '.jpg', '.jpeg', '.tiff', '.bmp', '.gif', '.webp')):
                try:
                    img = PILImage.open(input_path)
                    original_size = os.stat(input_path).st_size / 1024 / 1024
                    TOTAL_ORIGINAL += original_size

                    # Create output directory if it doesn't exist
                    os.makedirs(os.path.dirname(output_path), exist_ok=True)

                    # Convert PNG to JPG if required
                    if CONVERT_PNG_TO_JPG and file.lower().endswith('.png'):
                        img = img.convert('RGB')
                        output_path = output_path.replace('.png', '.jpg')

                    # Save the compressed image
                    img.save(output_path, optimize=True, quality=QUALITY)

                    compressed_size = os.stat(output_path).st_size / 1024 / 1024
                    TOTAL_COMPRESSED += compressed_size
                    gain = original_size - compressed_size
                    TOTAL_GAIN += gain
                    TOTAL_FILES += 1

                    compressed_files += 1
                    compression_progress = int((compressed_files / total_files) * 50)  # 0 to 50% progress for compression
                    progress_bar["value"] = compression_progress
                    progress_label.config(text=f"Compression Progress: {compression_progress}%")

                    print(f"Compressed: {input_path} -> {output_path}")
                    print(f"Original size: {original_size:.2f} MB, Compressed size: {compressed_size:.2f} MB, Gain: {gain:.2f} MB")

                except Exception as e:
                    print(f"Skipping file {input_path}: {e}")

            else:
                # Copy non-image files directly to the output folder
                os.makedirs(os.path.dirname(output_path), exist_ok=True)
                shutil.copy2(input_path, output_path)
                print(f"Copied non-image file: {input_path} -> {output_path}")

    # Final compression update
    progress_bar["value"] = 50  # Compression is complete (50% of the bar)
    progress_label.config(text="Compression completed. Proceeding to image insertion.")


# Function to handle subprocess output and update progress
def handle_subprocess_output(proc, progress_label, progress_bar):
    # Read and update progress from subprocess output
    for line in proc.stdout:
        if line:
            line = line.decode("utf-8").strip()
            if "Compressed:" in line:  # Customize to capture relevant lines for progress
                progress = extract_progress_from_line(line)
                progress_bar["value"] = progress
                progress_label.config(text=f"Progress: {progress}%")

    for error_line in proc.stderr:
        if error_line:
            error_line = error_line.decode("utf-8").strip()
            print(f"Subprocess Error: {error_line}")  # Log errors for debugging
    proc.wait()  # Wait for subprocess to finish


def extract_progress_from_line(line):
    # Example method to extract progress from the subprocess output
    # Adjust this based on the exact output of your subprocess
    if "Compressed:" in line:
        return 50  # Replace with actual logic to extract real progress
    return 0


# Main processing function
def process_images(queue, excel_file, image_folder, insert_col, product_col, should_compress, progress_label, progress_bar):
    if should_compress:
        output_folder = os.path.join(os.getcwd(), "output")
        os.makedirs(output_folder, exist_ok=True)

        # Run compression on images
        compress(image_folder, output_folder, progress_label, progress_bar)

    wb = load_workbook(excel_file)
    ws = wb.active

    processed_products = {}
    image_count = 0
    total_rows = ws.max_row - 1

    for row in range(2, ws.max_row + 1):
        product_code = ws[f"{product_col}{row}"].value

        if product_code and product_code not in processed_products:
            image_path = os.path.join(image_folder, f"{product_code}.jpg")

            if os.path.exists(image_path):
                cell_width = ws.column_dimensions[insert_col].width * 7.5
                cell_height = ws.row_dimensions[row].height or 250.8
                resized_img = resize_image_to_fit_cell(image_path, cell_width, cell_height)

                temp_image_path = os.path.join(image_folder, f"temp_{product_code}.jpg")
                resized_img.save(temp_image_path)

                img = ExcelImage(temp_image_path)
                img.width = resized_img.width
                img.height = resized_img.height
                ws.add_image(img, f"{insert_col}{row}")
                ws.row_dimensions[row].height = 250.8
                processed_products[product_code] = True
                image_count += 1
            else:
                ws[f"{insert_col}{row}"] = "No Image"
                ws.row_dimensions[row].height = 40
        elif product_code and product_code in processed_products:
            # Leave column A blank for repeated product codes and set row height to 40
            ws[f'A{row}'] = ""
            ws.row_dimensions[row].height = 40

        progress = int(50 + ((row - 1) / total_rows * 50))  # Progress bar will fill from 50 to 100 for image insertion
        progress_bar["value"] = progress
        progress_label.config(text=f"Progress: {progress}%")
        queue.put(progress)

    # Save the workbook
    save_path = os.path.splitext(excel_file)[0] + "_with_images.xlsx"
    wb.save(save_path)
    progress_label.config(text=f"Completed! Saved as: {os.path.basename(save_path)}")


# Function to start the threading process
def start_processing(queue, excel_file, image_folder, insert_col, product_col, should_compress, progress_label, progress_bar):
    threading.Thread(target=process_images, args=(queue, excel_file, image_folder, insert_col, product_col, should_compress, progress_label, progress_bar), daemon=True).start()


# GUI Setup
def main():
    root = tk.Tk()
    root.title("Excel Image Inserter")
    root.geometry("700x450")
    root.resizable(False, False)

    # Top Label
    top_label = tk.Label(root, text="Made by Artur Kuśmirek, All rights reserved", font=("Arial", 14, "bold"))
    top_label.pack(pady=10)

    # Input Fields and Buttons
    file_frame = tk.Frame(root)
    file_frame.pack(pady=20)

    tk.Label(file_frame, text="Excel File:").grid(row=0, column=0, padx=5, pady=5)
    excel_file_entry = tk.Entry(file_frame, width=40)
    excel_file_entry.grid(row=0, column=1, padx=5, pady=5)
    tk.Button(file_frame, text="Browse", command=lambda: excel_file_entry.insert(0, filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")]))).grid(row=0, column=2, padx=5, pady=5)

    tk.Label(file_frame, text="Image Folder:").grid(row=1, column=0, padx=5, pady=5)
    image_folder_entry = tk.Entry(file_frame, width=40)
    image_folder_entry.grid(row=1, column=1, padx=5, pady=5)
    tk.Button(file_frame, text="Browse", command=lambda: image_folder_entry.insert(0, filedialog.askdirectory())).grid(row=1, column=2, padx=5, pady=5)

    tk.Label(file_frame, text="Column to which you want to insert images:").grid(row=2, column=0, padx=5, pady=5)
    insert_col_entry = tk.Entry(file_frame, width=5)
    insert_col_entry.grid(row=2, column=1, sticky="w", padx=5)

    tk.Label(file_frame, text="Column with product codes:").grid(row=3, column=0, padx=5, pady=5)
    product_col_entry = tk.Entry(file_frame, width=5)
    product_col_entry.grid(row=3, column=1, padx=5, pady=5)

    should_compress_var = tk.BooleanVar()
    compress_check = tk.Checkbutton(file_frame, text="Compress Images", variable=should_compress_var)
    compress_check.grid(row=4, column=1, padx=5, pady=5)

    # Progress Bar
    progress_frame = tk.Frame(root)
    progress_frame.pack(pady=20)

    progress_label = tk.Label(progress_frame, text="Progress: 0%", font=("Arial", 12))
    progress_label.pack()

    progress_bar = ttk.Progressbar(progress_frame, length=500, mode="determinate")
    progress_bar.pack()

    # Start Processing Button
    def start_button_clicked():
        excel_file = excel_file_entry.get()
        image_folder = image_folder_entry.get()
        insert_col = insert_col_entry.get()
        product_col = product_col_entry.get()
        should_compress = should_compress_var.get()

        queue = Queue()
        start_processing(queue, excel_file, image_folder, insert_col, product_col, should_compress, progress_label, progress_bar)

    start_button = tk.Button(root, text="Start", command=start_button_clicked)
    start_button.pack(pady=10)

    root.mainloop()

if __name__ == "__main__":
    main()
